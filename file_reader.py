"""File reading helpers for txt / xls / xlsx score files."""

from pathlib import Path


def _fmt_cell(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:.10g}"
    return str(v)


def read_txt(path: str):
    with open(path, "r", encoding="utf-8") as f:
        lines = [l.strip() for l in f if l.strip()]
    ps, qm = [], []
    for line in lines:
        parts = line.split()
        if len(parts) >= 2:
            ps.append(parts[0])
            qm.append(parts[1])
    return ps, qm


def read_excel_sheets(path: str):
    """Return ``{sheet_name: [[cell, ...], ...]}`` for xls/xlsx files."""
    ext = Path(path).suffix.lower()
    sheets = {}
    if ext == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except ImportError as e:
            raise RuntimeError("openpyxl missing") from e
        wb = openpyxl.load_workbook(path, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append([_fmt_cell(v) for v in row])
            sheets[name] = data
    elif ext == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError as e:
            raise RuntimeError("xlrd missing") from e
        wb = xlrd.open_workbook(path)
        for name in wb.sheet_names():
            ws = wb.sheet_by_name(name)
            data = []
            for r in range(ws.nrows):
                data.append([_fmt_cell(ws.cell_value(r, c)) for c in range(ws.ncols)])
            sheets[name] = data
    return sheets
